# importing the Libraries
import pandas as pd 
import numpy as np 
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime  
import dateutil.parser
import os
import warnings  

# filter any warnings so that it will not showcase any warnings in terminal
warnings.filterwarnings('ignore')

# creating some fake columns 
lst_of_fake_col = []
for i in range(1,100): 
    lst_of_fake_col.append(f'fake_col-{i}')  

# Get a list of all the file names in the folder
folder = input('Enter the folder path: ')
if folder[0] == '"' and folder[-1] == '"':
    folder = folder[1:]
    folder = folder[:-1]
    filenames = os.listdir(folder)
    

for file in filenames:
    filepath = os.path.join(folder, file)
    
    # reading the files, with fake columns as headers, and dropping all the rows and columns containing NaN
    df = pd.read_csv(filepath,names= lst_of_fake_col)  
    df.dropna(how = 'all',inplace=True)
    df.dropna(how='all', axis=1,inplace=True)  

    # rows containg '#=' [Adobe info data] 
    adobe_info =  df[df['fake_col-1'].str.contains("#=================================================================") == True] 

    # lst_of_index_having_hash_equal
    lst_of_index_having_hash_equal = list(adobe_info.index)  

    # to locate the date b/w the cols of adobe_info
    to_locate_the_date = df.iloc[lst_of_index_having_hash_equal[0]:lst_of_index_having_hash_equal[1]+1]   
    a = list(to_locate_the_date['fake_col-1']) 

    # extracting only date from the string 
    for i in a:
        if 'Date' in i:

            str_date = i[-12:].replace(',','')
            str_date1 = str_date.replace(' ','/')

    # Extracted_date = datetime.datetime.strptime(str_date1, "%b/%d/%Y").date()     # changing the format of string date to datetime.date

    Extracted_date = datetime.datetime.strptime(str_date1, "%b/%d/%Y").date()

    # droping all the rows which are having adobe information
    lst1 = lst_of_index_having_hash_equal[::2]
    lst2 = lst_of_index_having_hash_equal[1::2] 

    # its the index for jan-feb-march etc ( it's the index of the table header)
    index_of_Funnel_trendviews = list(df[pd.notna(df['fake_col-1']) & df['fake_col-1'].str.contains('# Funnel_Trend Views')].index)  
    index_of_content_type_trendviews = list(df[pd.notna(df['fake_col-1']) & df['fake_col-1'].str.contains('# Content Type_Trend View')].index)  
    index_of_all_regions_trendviews = list(df[pd.notna(df['fake_col-1']) & df['fake_col-1'].str.contains('# All Regions_Trend Views')].index)

    # the above will not give the correct index no. we need to add 
    header_of_Funnel_trendviews = index_of_Funnel_trendviews[0] +2 
    header_of_content_type_trendviews = index_of_content_type_trendviews[0] + 2 
    # header_of_Global_trendviews = index_of_Global_trendviews[0] + 2 
    header_of_all_regions_trendviews = index_of_all_regions_trendviews[0]  + 2     

    # for the above index getting it's respective rows and setting it to the dataframes
    df.iloc[[header_of_all_regions_trendviews]]
    unpivot0_header_list = df.iloc[[header_of_all_regions_trendviews]].values.tolist()[0] 

    df.iloc[[header_of_Funnel_trendviews]]  
    unpivot1_header_list = df.iloc[[header_of_Funnel_trendviews]].values.tolist()[0] 

    df.iloc[[header_of_content_type_trendviews]] 
    unpivot2_header_list = df.iloc[[header_of_content_type_trendviews]].values.tolist()[0] 

    # dropping all the rows which contains the adobe info
    for r in range(len(lst1)):
        df.drop(df.loc[lst1[r]:lst2[r]].index,axis=0,inplace=True) 

    # to remove all the values in the table having #####
    df1 =  df[df["fake_col-1"].str.contains("##############################################") == False] 

    # converting the date to str date
    from datetime import date

    extracted_date = str(Extracted_date) 
    year, month, _ = map(int, extracted_date.split("-"))
    month_year = date(year, month, 1)
    month_year_str = month_year.strftime("%b-%Y")
    month_year_str = month_year_str.replace("-"," ")

    # drop old index, refresh it and inplace it into the original df1
    df1.reset_index(drop=True,inplace=True)

    # now getting the rows index which starts with '#' (they are the tables naemes) 
    df2 = df1[df1['fake_col-1'].str.startswith('#')]  

    # All tables names
    all_tables_names = list(df2['fake_col-1']) 

    # list of index having column headers 
    lst_of_index_having_column_header = list(df2[df2['fake_col-1'].str.startswith('#')].index)

    # total no of tables which would be created is nothing but the len of list_of_index_having_column_header
    a = len(lst_of_index_having_column_header) 
    print(f'total no. of tables in the file at {file} => {a}')            #<-------------------------------------

    # lst1 will contain all the columns header index
    ls1 = lst_of_index_having_column_header 

    # lst2 will contain all the columns header index + 2
    ls2= []
    for i in lst_of_index_having_column_header :
        ls2.append(i+2) 

    tables_names = []
    for i in all_tables_names:
        tables_names.append((i)) 
    # print(tables_names)

    # to store the dataframes with respect to its keys in a dict.
    k = len(tables_names) 
    tables = {}
    for i in range(0,k):
        if i<k-1: 
            tables[tables_names[i]] = df1.iloc[ls2[i]:ls1[i+1]]

        elif i==k-1:
            tables[tables_names[i]] = df1.iloc[ls2[i]:] 


    # date required tables mai date date daalege, country required table mai country daalenge and global required mai global dalenge
    date_required_tables = ['# All Regions', '# All Metrics by Month', '# Managed Vs Unmanaged', '# Funnel', '# Content Type', '# Node Type','# Global', '# US', '# UK', '# DE', '# FR', '# JP', '# CA', '# PRC', '# IN', '# AU']
    country_required_tables = ['# US', '# UK', '# DE', '# FR', '# JP', '# CA', '# PRC', '# IN', '# AU']
    global_coutries_tables = ['# Global'] 

    for i in list(tables.keys()):
        tables[i].replace('Infinity', 0, inplace=True)

        if i in date_required_tables:
            tables[i].insert(0, "Date", Extracted_date)


        if i in country_required_tables:
            tables[i].insert(1,'Country',i[2:])


        if i in global_coutries_tables: 
            tables[i].insert(1, "Country", 'Global')

    # inserting the attribute column to the table of the panel 1
    tables['# All Regions'].insert(1, 'Attribute', 'All Regions')
    tables['# All Metrics by Month'].insert(1, 'Attribute', 'All Metrics by Month')
    tables['# Funnel'].insert(1, 'Attribute', 'Funnel')
    tables['# Content Type'].insert(1, 'Attribute', 'Content Type')
    tables['# Node Type'].insert(1, 'Attribute', 'Node Type') 
    tables['# Managed Vs Unmanaged'].insert(1,'Attribute','Managed Vs Unmanaged') 

    # sheet1 table
    sheet1 = ['# All Regions', '# All Metrics by Month', '# Managed Vs Unmanaged', '# Funnel', '# Content Type', '# Node Type']

    # iterate over the all_tables_name and if the tables are in the sheet1 then append to list:- list_of_tables_names_for_sheet1
    list_of_tables_names_for_sheet1 = []
    for i in all_tables_names:  # fixed index out of range error
        if i in sheet1:
            list_of_tables_names_for_sheet1.append(tables[i])  

    # metric summary - month level ke table mai sheet 1 ke saare tables append kr rhe h
    tables['Metric Summary - Month Level'] = list_of_tables_names_for_sheet1[0].append([list_of_tables_names_for_sheet1[i] for i in range(1,len(list_of_tables_names_for_sheet1))])

    # all_regions tables
    all_regions = ['# Global','# US', '# UK', '# DE', '# FR', '# JP', '# CA', '# PRC', '# IN', '# AU']

    # all the regions tables ko apeend in a single table called Page_URLs
    list_of_tables_names_for_coutries_tables = []
    for i in all_tables_names:  # fixed index out of range error
        if i in all_regions:
            list_of_tables_names_for_coutries_tables.append(tables[i])  

    tables['Page_URLs'] = list_of_tables_names_for_coutries_tables[0].append([list_of_tables_names_for_coutries_tables[i] for i in range(1,len(list_of_tables_names_for_coutries_tables))]) 

    # resetting the index
    tables['Metric Summary - Month Level'].reset_index(drop=True,inplace=True) 
    tables['Page_URLs'].reset_index(drop=True,inplace=True) 

    # deleting all the tables from tables dict having the below keys
    del_tables_keys = sheet1 + all_regions 

    for i in del_tables_keys:
        if i in tables.keys():
            del tables[i]   

    # all regions trend views and global trend views ke table ki date index ko string se change kr ke datetime mai convert kr do.
    for i in list(tables.keys()):
        if i in ['# All Regions_Trend Views','# Global_Trend View']: 
            tables[i].iloc[:,0] = pd.to_datetime(tables[i].iloc[:,0],infer_datetime_format=True, errors='ignore') 

    # unpivot header list ka 1st index should be None becoz if we table we need a none col header to unpivot the dataframe
    unpivot1_header_list[0] = None 
    unpivot2_header_list[0] = None 
    unpivot0_header_list[0] = None 

    # setting up the columns header to the tables which need to be unpivot
    tables['# Funnel_Trend Views'].columns = unpivot1_header_list 
    tables['# Content Type_Trend View'].columns = unpivot2_header_list 
    tables['# All Regions_Trend Views'].columns = unpivot0_header_list  

    # unpivoting the tables:- id_vars mai voh column aata hai jisko repeated way mai column se row mai transpose krna hai
    melted0 = pd.melt(tables['# All Regions_Trend Views'] , id_vars= [None], value_vars= unpivot0_header_list)  
    melted1 = pd.melt(tables['# Funnel_Trend Views'] , id_vars= [None], value_vars= unpivot1_header_list)   
    melted2 = pd.melt(tables['# Content Type_Trend View'] , id_vars= [None], value_vars= unpivot2_header_list)     

    # now giving refrence to the melted tables to the dict tables
    tables['# All Regions_Trend Views'] = melted0
    tables['# Funnel_Trend Views'] = melted1
    tables['# Content Type_Trend View'] = melted2

    # inserting the default attribute column to the Trend views panel tables
    tables['# Funnel_Trend Views'].insert(0, 'Attribute', 'Funnel_Trend Views')
    tables['# Content Type_Trend View'].insert(0, 'Attribute', 'Content Type_Trend View')
    tables['# All Regions_Trend Views'].insert(1, 'Attribute', 'All Regions_Trend Views') 
    tables['# Global_Trend View'].insert(1, 'Attribute', 'Region')
    tables['# Global_Trend View'].insert(2, 'Values', 'Global') 


    # giving the sheet names 
    writer = pd.ExcelWriter('file-without-#.xlsx', engine = 'xlsxwriter')

    for nn in list(tables.keys()):

        try:
            if nn == 'Metric Summary - Month Level':
                if (tables[nn].iloc[0,0] == Extracted_date and tables[nn].iloc[0,3].isdigit()):
    #                 print(nn)
                    tables[nn].to_excel(writer, sheet_name = nn,header = False,index=False)

            elif nn == 'Page_URLs':
                if tables[nn].iloc[0,0] == Extracted_date  and tables[nn].iloc[0,1] in [i[2:] for i in all_regions]:
    #                 print(nn)
                    tables[nn].to_excel(writer, sheet_name = 'Page_URLs',header = False,index=False)  

            elif nn == '# All Regions_Trend Views':
                if isinstance(tables[nn].iloc[0,0], datetime.date) or tables[nn].iloc[0,3].isdigit():
    #                 print(nn)
                    tables[nn].to_excel(writer, sheet_name = nn[2:] ,header = False,index=False) 

            elif nn == '# Funnel_Trend Views':
                if tables[nn].iloc[0,1][0:10] == 'CSE Funnel' or tables[nn].iloc[0,3].isdigit():
    #                 print(nn)
                    tables[nn].to_excel(writer, sheet_name = nn[2:], header = False,index=False)

            elif nn == '# Content Type_Trend View':
    #             print(nn) 
                if tables[nn].iloc[0,1][0:16] == 'CCM Content Type' or tables[nn].iloc[0,3].isdigit():
    #                 print(nn)
                    tables[nn].to_excel(writer, sheet_name = nn[2:], header = False,index=False)

            elif nn == '# Global_Trend View':
                if isinstance(tables[nn].iloc[0,0], datetime.date)  or tables[nn].iloc[0,3].isdigit():
    #                 print(nn)
                    tables[nn].to_excel(writer, sheet_name = nn[2:] ,header = False,index=False)    
            else:
                #print(nn)
                tables[nn].to_excel(writer, sheet_name =f'UDF' ,header = False,index=False)

        except Exception:
            pass 

    writer.close() 

    # giving the custome header to each workbook:-
    import openpyxl 

    wb = openpyxl.load_workbook('file-without-#.xlsx') 

    sheetnames_lst = []
    for i in wb.sheetnames:
        sheetnames_lst.append(i)

#     print(sheetnames_lst) 

    columns_for_all_regions = ['Date','Attribute','Values','Page Views','Unique Visitors','Single Page Visits','Exits','Time Spent per Visit (seconds)','Page Content Interaction Rate','Scroll Rate: 25%','Scroll Rate: 50%','Scroll Rate: 75%','Scroll Rate: 100%']
    # columns_for_MvsU = ['Date','Segment','Page Views','Single Page Visits','Exits','Time Spent per Visit (seconds)','Page Content Interaction Rate','Scroll Rate: 25%','Scroll Rate: 50%','Scroll Rate: 75%','Scroll Rate: 100%']
    columns_for_df_concat = ['Date','Country','Page URL (evar35)','Page Views','Unique Visitors','Single Page Visits','Exits','Time Spent per Visit (seconds)','Page Content Interaction Rate','Scroll Rate: 25%','Scroll Rate: 50%','Scroll Rate: 75%','Scroll Rate: 100%']
    columns_for_AllRegions_ = ['Month','Attribute','Values','Page Views']
    columns_for_Global_ = ['Month','Attribute','Values','Page Views','Single Page Visits','Exits']
    columns_for_ContentType_funnel_ = ['Attribute','Values','Month','Page Views']


    for i in sheetnames_lst:
        ws = wb[i]  
        ws.insert_rows(1,1)

        if i == 'Metric Summary - Month Level':
            for j in range(0,len(columns_for_all_regions)):
                ws.cell(1,j+1).value = columns_for_all_regions[j] 

        elif i == 'Page_URLs':
            for l in range(0,len(columns_for_df_concat)):
                ws.cell(1,l+1).value = columns_for_df_concat[l]

        elif i == 'All Regions_Trend Views':
            for zz in range(0,len(columns_for_AllRegions_)):
                ws.cell(1,zz+1).value = columns_for_AllRegions_[zz] 

        elif i in ['Funnel_Trend Views','Content Type_Trend View']:
            for m in range(0,len(columns_for_ContentType_funnel_)):
                ws.cell(1,m+1).value = columns_for_ContentType_funnel_[m] 

        elif i == 'Global_Trend View':
            for n in range(0,len(columns_for_Global_)):
                ws.cell(1,n+1).value = columns_for_Global_[n]

        else:
            print('------------------------------------------------\n')
            print(f'Header not Found For Sheet {i}')
            print('------------------------------------------------\n')  

        # Create the folder path
        folder_path = 'Single File Analysis'

        # Check if the folder already exists
        if not os.path.exists(folder_path):
            # If the folder does not exist, create it
            os.makedirs(folder_path)

        # Save the file with the extracted date as part of the file name
        file_path = os.path.join(folder_path, f'Digital_Strategy_AA - {Extracted_date}.xlsx')

        # Check if the file already exists
        if os.path.exists(file_path):
            # If the file exists, delete it
            os.remove(file_path)

        # Save the workbook
        wb.save(file_path)    

print(f'\nAll Files Are Parsed Successfully\n')

os.remove('file-without-#.xlsx') 

print('\nDo you want to Create a Master File? (Yes or No)')
Master_file_input = input()

if Master_file_input in ['Yes','yes']:

    # Create the folder path
    Master_Folder = 'Master File Analysis'

    # Check if the folder already exists
    if not os.path.exists(Master_Folder):
        # If the folder does not exist, create it
        os.makedirs(Master_Folder)
        
    folder1 = 'Single File Analysis'
    filenames1 = os.listdir(folder1)

    # Create a new Excel workbook
    workbook = None
    for file in filenames1:
        filepath1 = os.path.join(folder1, file)
        xl = pd.ExcelFile(filepath1)
        empty_flag = False
        # Iterate over the sheet names
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name)
            if df.empty:
                empty_flag = True
                continue  # Skip the empty sheet
            if not empty_flag:
                if workbook is None:
                    workbook = openpyxl.Workbook(write_only=True)
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    sheet = workbook.create_sheet(title=sheet_name)
                    sheet.append(df.columns.tolist())
                for r in df.iterrows():
                    sheet.append(r[1].tolist())
        if empty_flag:
            print(f'{file} has empty sheet.')
    # Save the workbook to the Master File Analysis folder
    workbook.save(os.path.join(Master_Folder, 'Digital_Strategy_AA - Master.xlsx'))
    print(f'\nMaster File Created Successfully\n') 
